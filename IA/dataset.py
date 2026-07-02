"""
IA/dataset.py — Chargement de datasets multi-formats.

Formats supportes (detection automatique par extension) :
  - CSV, TSV        : csv/tsv (standard Python)
  - JSON, JSONL     : json (standard Python)
  - NPY, NPZ        : numpy
  - TXT             : texte brut (mots/lignes)
  - Parquet         : pyarrow (optionnel)
  - HDF5 / H5       : h5py (optionnel)
  - XLSX / XLS      : openpyxl (optionnel)
  - Pickle / PKL    : pickle (standard Python)

Usage :
    from IA.dataset import load_dataset
    X, y = load_dataset("donnees.csv", target="label")
    X, y = load_dataset("donnees.json", target="classe")
    data = load_dataset("features.npy")  # retourne X seul si pas de target

Convention :
  - Les colonnes numeriques deviennent les features (X).
  - La colonne "target" specifiee devient le label (y).
  - Si target=None, retourne (data, None).
  - Toutes les donnees sont converties en float64 numpy arrays.
"""

import os
import csv
import json
import logging
import struct
import io
import pickle
from typing import Optional, Tuple, Union, List, Dict, Any

import numpy as np

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Extensions mappees vers des handlers
# ---------------------------------------------------------------------------

_LOADERS = {}


def _register(ext):
    """Decorateur pour enregistrer un loader par extension."""
    def deco(func):
        for e in ext if isinstance(ext, (list, tuple, set, frozenset)) else [ext]:
            _LOADERS[e.lower()] = func
        return func
    return deco


# ============================================================================
# CSV / TSV
# ============================================================================

@_register([".csv", ".tsv"])
def _load_csv(path, target=None, **kw):
    """Charge un CSV/TSV. Detection auto du delimiter par extension."""
    delim = "\t" if path.lower().endswith(".tsv") else ","
    encoding = kw.get("encoding", "utf-8")

    with open(path, "r", encoding=encoding, newline="") as f:
        # Detecter si la premiere ligne est un header (verifie que les
        # valeurs de la premiere ligne ne sont pas toutes numeriques)
        sample = f.read(8192)
        f.seek(0)

        reader = csv.reader(f, delimiter=delim)
        rows = list(reader)

    if not rows:
        raise ValueError(f"Fichier vide : {path}")

    # Detecter header
    has_header = _has_header(rows)

    if has_header:
        header = rows[0]
        data_rows = rows[1:]
        if not data_rows:
            raise ValueError(f"CSV avec header mais aucune donnee : {path}")

        # Trouver les indices des colonnes
        target_idx = None
        feature_indices = []
        for i, col in enumerate(header):
            col_stripped = col.strip()
            if target and col_stripped.lower() == target.lower():
                target_idx = i
            else:
                feature_indices.append(i)
    else:
        header = None
        data_rows = rows
        # Si pas de header et target specifie, on suppose que
        # la derniere colonne est le target
        if target is not None:
            raise ValueError("target specifie mais le CSV n'a pas de header")
        ncols = len(data_rows[0])
        feature_indices = list(range(ncols - 1))
        target_idx = ncols - 1

    # Parser les donnees
    X_rows, y_rows = [], []
    for row in data_rows:
        if not row:
            continue
        feat_vals = []
        for i in feature_indices:
            if i < len(row):
                feat_vals.append(_parse_num(row[i]))
            else:
                feat_vals.append(0.0)
        X_rows.append(feat_vals)

        if target_idx is not None and target_idx < len(row):
            y_rows.append(_parse_num(row[target_idx]))

    X = np.array(X_rows, dtype=np.float64)
    y = np.array(y_rows, dtype=np.float64).reshape(-1, 1) if y_rows else None

    return X, y, header


# ============================================================================
# JSON / JSONL
# ============================================================================

@_register([".json", ".jsonl", ".ndjson"])
def _load_json(path, target=None, **kw):
    """Charge un JSON (liste d'objets) ou JSONL (un objet par ligne)."""
    with open(path, "r", encoding=kw.get("encoding", "utf-8")) as f:
        content = f.read()

    if path.lower().endswith((".jsonl", ".ndjson")):
        # JSON Lines: un objet par ligne
        records = [json.loads(line) for line in content.strip().splitlines() if line.strip()]
    else:
        data = json.loads(content)
        if isinstance(data, list):
            records = data
        elif isinstance(data, dict):
            # Peut-etre {"data": [...], "labels": [...]} ou similaire
            for key in ("data", "X", "features", "samples", "inputs"):
                if key in data:
                    records = data[key]
                    # Chercher les labels separement
                    y_key = None
                    for yk in ("labels", "y", "targets", "classes", target or ""):
                        if yk and yk in data:
                            y_key = yk
                            break
                    if y_key:
                        y_arr = np.array(data[y_key], dtype=np.float64)
                        if y_arr.ndim == 1:
                            y_arr = y_arr.reshape(-1, 1)
                        X_arr = _to_numeric_array(records)
                        return X_arr, y_arr, list(records[0].keys()) if isinstance(records[0], dict) else None
                    break
            else:
                records = [data]
        else:
            records = [data]

    if not records:
        raise ValueError(f"JSON vide ou invalide : {path}")

    if isinstance(records[0], dict):
        # Liste de dictionnaires — extraire features et target
        keys = list(records[0].keys())
        target_key = None
        feature_keys = []
        for k in keys:
            if target and k.lower() == target.lower():
                target_key = k
            else:
                feature_keys.append(k)

        if target_key is None and target is not None:
            raise ValueError(f"Cle target '{target}' non trouvee. Cles: {keys}")

        X_rows, y_rows = [], []
        for rec in records:
            feat = [_parse_num(rec.get(k, 0)) for k in feature_keys]
            X_rows.append(feat)
            if target_key is not None:
                y_rows.append(_parse_num(rec.get(target_key, 0)))

        X = np.array(X_rows, dtype=np.float64)
        y = np.array(y_rows, dtype=np.float64).reshape(-1, 1) if y_rows else None
        return X, y, keys
    else:
        # Liste de listes/numbers
        X = _to_numeric_array(records)
        return X, None, None


# ============================================================================
# NPY / NPZ
# ============================================================================

@_register([".npy"])
def _load_npy(path, target=None, **kw):
    """Charge un fichier .npy (numpy array)."""
    arr = np.load(path)
    arr = arr.astype(np.float64)
    return arr, None, None


@_register([".npz"])
def _load_npz(path, target=None, **kw):
    """Charge un fichier .npz (numpy zipped archive)."""
    data = np.load(path, allow_pickle=True)
    keys = list(data.keys())

    # Trouver X et y par convention de nommage
    x_key = y_key = None
    for k in keys:
        kl = k.lower()
        if kl in ("x", "features", "data", "inputs", "samples"):
            x_key = k
        elif kl in ("y", "labels", "targets", "classes"):
            y_key = k
        elif target and kl == target.lower():
            y_key = k

    if x_key is None:
        # Prendre le premier
        x_key = keys[0]
        if len(keys) > 1:
            y_key = keys[1]

    X = data[x_key].astype(np.float64)
    y = data[y_key].astype(np.float64).reshape(-1, 1) if y_key else None
    return X, y, keys


# ============================================================================
# TXT (texte brut)
# ============================================================================

@_register([".txt"])
def _load_txt(path, target=None, **kw):
    """Charge un fichier texte.
    - Si toutes les lignes sont numeriques: traite comme CSV sans header.
    - Sinon: retourne les lignes brutes (pour NLP/SLM).
    """
    encoding = kw.get("encoding", "utf-8")
    with open(path, "r", encoding=encoding) as f:
        lines = [l.strip() for l in f if l.strip()]

    if not lines:
        raise ValueError(f"Fichier texte vide : {path}")

    # Verifier si c'est purement numerique
    all_numeric = all(_is_numeric_line(l) for l in lines)

    if all_numeric:
        # Parser comme donnees numeriques (delimiter auto: espace, virgule, tab)
        parsed = []
        for line in lines:
            # Essayer differents delimiteurs
            for delim in [",", "\t", ";", " "]:
                parts = line.split(delim)
                if len(parts) > 1:
                    parsed.append([_parse_num(p) for p in parts])
                    break
            else:
                parsed.append([_parse_num(line)])

        arr = np.array(parsed, dtype=np.float64)
        if target and arr.shape[1] > 1:
            # target est l'index de colonne
            tidx = int(target) if target.isdigit() else arr.shape[1] - 1
            return arr[:, :tidx], arr[:, tidx:tidx+1], None
        return arr, None, None
    else:
        # Texte brut — retourne les lignes telles quelles
        # Utile pour SLM, NLP, etc.
        return np.array(lines, dtype=object), None, None


# ============================================================================
# HDF5 / H5
# ============================================================================

@_register([".h5", ".hdf5"])
def _load_hdf5(path, target=None, **kw):
    """Charge un fichier HDF5. Necessite h5py."""
    try:
        import h5py
    except ImportError:
        raise ImportError(
            "h5py requis pour lire les fichiers HDF5. "
            "Installez avec: pip install h5py"
        )

    with h5py.File(path, "r") as f:
        keys = list(f.keys())

        x_key = y_key = None
        for k in keys:
            kl = k.lower()
            if kl in ("x", "features", "data", "inputs"):
                x_key = k
            elif kl in ("y", "labels", "targets", "classes"):
                y_key = k
            elif target and kl == target.lower():
                y_key = k

        if x_key is None:
            x_key = keys[0]
            if len(keys) > 1:
                y_key = keys[1]

        X = np.array(f[x_key], dtype=np.float64)
        y = np.array(f[y_key], dtype=np.float64).reshape(-1, 1) if y_key else None

    return X, y, keys


# ============================================================================
# Parquet
# ============================================================================

@_register([".parquet", ".pq"])
def _load_parquet(path, target=None, **kw):
    """Charge un fichier Parquet. Necessite pyarrow."""
    try:
        import pyarrow.parquet as pq
    except ImportError:
        raise ImportError(
            "pyarrow requis pour lire les fichiers Parquet. "
            "Installez avec: pip install pyarrow"
        )

    table = pq.read_table(path)
    df = table.to_pydict()

    keys = list(df.keys())
    target_key = None
    feature_keys = []

    for k in keys:
        if target and k.lower() == target.lower():
            target_key = k
        else:
            feature_keys.append(k)

    X_cols = []
    for k in feature_keys:
        col = df[k]
        if hasattr(col, '__len__'):
            X_cols.append(np.array(col, dtype=np.float64))
        else:
            X_cols.append(np.full(len(next(iter(df.values()))), float(col), dtype=np.float64))

    X = np.column_stack(X_cols) if X_cols else np.empty((0, 0), dtype=np.float64)

    y = None
    if target_key:
        y_arr = np.array(df[target_key], dtype=np.float64).reshape(-1, 1)
        y = y_arr

    return X, y, keys


# ============================================================================
# XLSX / XLS (Excel)
# ============================================================================

@_register([".xlsx", ".xls"])
def _load_excel(path, target=None, **kw):
    """Charge un fichier Excel. Necessite openpyxl (.xlsx) ou xlrd (.xls)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "openpyxl requis pour lire les fichiers Excel. "
            "Installez avec: pip install openpyxl"
        )

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError(f"Fichier Excel vide : {path}")

    # Detecter header
    has_header = _has_header(rows)

    if has_header:
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        data_rows = rows[1:]
    else:
        header = None
        data_rows = rows

    if not data_rows:
        raise ValueError(f"Excel sans donnees : {path}")

    ncols = len(data_rows[0])
    feature_indices = list(range(ncols))
    target_idx = None

    if header and target:
        for i, col in enumerate(header):
            if col.lower() == target.lower():
                target_idx = i
                feature_indices = [j for j in range(ncols) if j != i]
                break

    X_rows, y_rows = [], []
    for row in data_rows:
        feat = [_parse_num(row[i] if i < len(row) else 0) for i in feature_indices]
        X_rows.append(feat)
        if target_idx is not None:
            y_rows.append(_parse_num(row[target_idx] if target_idx < len(row) else 0))

    X = np.array(X_rows, dtype=np.float64)
    y = np.array(y_rows, dtype=np.float64).reshape(-1, 1) if y_rows else None

    return X, y, header


# ============================================================================
# Pickle / PKL
# ============================================================================

@_register([".pkl", ".pickle", ".pkl.gz"])
def _load_pickle(path, target=None, **kw):
    """Charge un fichier pickle. Accepte np.array, dict, tuple (X, y), ou liste."""
    open_func = gzip.open if path.endswith(".gz") else open
    with open_func(path, "rb") as f:
        data = pickle.load(f)

    if isinstance(data, tuple) and len(data) == 2:
        X = np.array(data[0], dtype=np.float64)
        y = np.array(data[1], dtype=np.float64).reshape(-1, 1)
        return X, y, None
    elif isinstance(data, dict):
        keys = list(data.keys())
        x_key = y_key = None
        for k in keys:
            kl = str(k).lower()
            if kl in ("x", "features", "data", "inputs"):
                x_key = k
            elif kl in ("y", "labels", "targets", "classes"):
                y_key = k
            elif target and kl == target.lower():
                y_key = k
        if x_key is None:
            x_key = keys[0]
            if len(keys) > 1:
                y_key = keys[1]
        X = np.array(data[x_key], dtype=np.float64)
        y = np.array(data[y_key], dtype=np.float64).reshape(-1, 1) if y_key else None
        return X, y, keys
    else:
        return np.array(data, dtype=np.float64), None, None


# ============================================================================
# GZIP (delegue au bon loader apres decompression)
# ============================================================================

try:
    import gzip
    _HAS_GZIP = True
except ImportError:
    _HAS_GZIP = False


@_register([".gz"])
def _load_gzip(path, target=None, **kw):
    """Charge un fichier gzip. Detecte le format interne par extension."""
    if not _HAS_GZIP:
        raise ImportError("gzip non disponible")

    # Trouver l'extension reelle
    base = path[:-3]  # retire .gz
    _, ext = os.path.splitext(base)
    if ext.lower() in _LOADERS:
        # Decompresser dans un StringIO/BytesIO
        with gzip.open(path, "rb") as gz:
            content = gz.read()

        # Ecrire dans un temp file
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name

        try:
            return _LOADERS[ext.lower()](tmp_path, target=target, **kw)
        finally:
            os.unlink(tmp_path)
    else:
        raise ValueError(f"Format gzip non reconnu (extension interne: {ext})")


# ============================================================================
# Fonctions utilitaires
# ============================================================================

def _parse_num(val) -> float:
    """Convertit une valeur en float. Tolere les int, float, str."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float, np.integer, np.floating)):
        return float(val)
    s = str(val).strip()
    # Gerer les formats avec espaces ou guillemets
    s = s.strip('"\'').strip()
    try:
        return float(s)
    except ValueError:
        # Essayer de remplacer la virgule decimale
        s = s.replace(",", ".")
        try:
            return float(s)
        except ValueError:
            logger.warning("Valeur non numerique ignoree: %r -> 0.0", val)
            return 0.0


def _is_numeric_line(line: str) -> bool:
    """Verifie si une ligne de texte est purement numerique."""
    for part in line.replace(",", " ").replace("\t", " ").split():
        try:
            float(part)
        except ValueError:
            return False
    return True


def _has_header(rows) -> bool:
    """Detecte si la premiere ligne est un header (contient du texte)."""
    if len(rows) < 2:
        return False
    first_row = rows[0]
    # Si toutes les valeurs de la premiere ligne sont numeriques, pas de header
    for val in first_row:
        try:
            float(str(val).strip().strip('"\''))
        except (ValueError, TypeError):
            return True  # Au moins une valeur non numerique = header
    return False


def _to_numeric_array(data) -> np.ndarray:
    """Convertit une liste de listes/nombres en numpy array float64."""
    if isinstance(data, np.ndarray):
        return data.astype(np.float64)
    if isinstance(data, list) and data and isinstance(data[0], (list, tuple)):
        return np.array(data, dtype=np.float64)
    return np.array(data, dtype=np.float64).reshape(-1, 1)


# ============================================================================
# API publique
# ============================================================================

def load_dataset(
    path: str,
    target: Optional[str] = None,
    **kwargs
) -> Union[Tuple[np.ndarray, np.ndarray],
           Tuple[np.ndarray, None],
           np.ndarray]:
    """Charge un dataset depuis un fichier.

    Detection automatique du format par extension de fichier.

    Args:
        path: Chemin vers le fichier de donnees.
        target: Nom de la colonne/cle a utiliser comme label (y).
                Si None, retourne (X, None).
        **kwargs: Arguments supplementaires passes au loader specifique.
                  - encoding: encodage du texte (defaut: utf-8)
                  - delimiter: delimiter CSV (auto-detecte par defaut)

    Returns:
        (X, y) ou X si y est None.
        X: numpy array float64 de shape (n_samples, n_features).
        y: numpy array float64 de shape (n_samples, 1) ou None.

    Raises:
        ValueError: fichier vide ou format non reconnu.
        ImportError: dependance optionnelle manquante (h5py, pyarrow, etc.)

    Examples:
        >>> X, y = load_dataset("data.csv", target="label")
        >>> X, y = load_dataset("data.json", target="classe")
        >>> X = load_dataset("features.npy")
        >>> X, y = load_dataset("data.parquet", target="target")
        >>> X, y = load_dataset("data.h5")
        >>> lines, _ = load_dataset("texte.txt")  # texte brut

    Formats supportes:
        .csv, .tsv      — CSV/TSV (standard, zero dep)
        .json, .jsonl   — JSON/JSON Lines (standard, zero dep)
        .npy, .npz      — NumPy (numpy requis)
        .txt            — Texte brut ou numerique (standard, zero dep)
        .h5, .hdf5      — HDF5 (h5py requis)
        .parquet, .pq   — Parquet (pyarrow requis)
        .xlsx, .xls     — Excel (openpyxl requis)
        .pkl, .pickle   — Pickle (standard, zero dep)
        .gz             — GZip (auto-delegue, standard)
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"Fichier non trouve : {path}")

    _, ext = os.path.splitext(path)
    ext_lower = ext.lower()

    # Cas special: .pkl.gz
    if path.endswith(".pkl.gz"):
        ext_lower = ".pkl.gz"

    loader = _LOADERS.get(ext_lower)
    if loader is None:
        supported = ", ".join(sorted(_LOADERS.keys()))
        raise ValueError(
            f"Format non supporte : {ext}. Formats supportes : {supported}"
        )

    logger.info("Chargement dataset : %s (format: %s, target: %s)",
                path, ext_lower, target)

    X, y, header = loader(path, target=target, **kwargs)

    logger.info("Dataset charge : X=%s, y=%s",
                X.shape, y.shape if y is not None else "None")

    return (X, y) if y is not None else X


def supported_formats() -> List[str]:
    """Retourne la liste des formats supportes."""
    return sorted(_LOADERS.keys())


def dataset_info(path: str, **kwargs) -> Dict[str, Any]:
    """Retourne des informations sur un dataset sans le charger entierement.

    Returns:
        dict avec clefs: format, size_bytes, n_rows, n_cols, has_target, columns
    """
    size = os.path.getsize(path)
    _, ext = os.path.splitext(path)
    ext_lower = ext.lower()

    if ext_lower in (".csv", ".tsv"):
        delim = "\t" if ext_lower == ".tsv" else ","
        encoding = kwargs.get("encoding", "utf-8")
        with open(path, "r", encoding=encoding, newline="") as f:
            reader = csv.reader(f, delimiter=delim)
            rows = list(reader)
        has_header = _has_header(rows)
        header = rows[0] if has_header else None
        data_rows = rows[1:] if has_header else rows
        return {
            "format": ext_lower,
            "size_bytes": size,
            "n_rows": len(data_rows),
            "n_cols": len(data_rows[0]) if data_rows else 0,
            "has_header": has_header,
            "columns": header,
        }
    elif ext_lower in (".json", ".jsonl", ".ndjson"):
        with open(path, "r") as f:
            if ext_lower in (".jsonl", ".ndjson"):
                records = [json.loads(l) for l in f if l.strip()]
            else:
                data = json.load(f)
                records = data if isinstance(data, list) else [data]
        n_rows = len(records)
        n_cols = len(records[0]) if records and isinstance(records[0], (dict, list)) else 0
        keys = list(records[0].keys()) if records and isinstance(records[0], dict) else None
        return {
            "format": ext_lower,
            "size_bytes": size,
            "n_rows": n_rows,
            "n_cols": n_cols,
            "columns": keys,
        }
    elif ext_lower in (".npy",):
        arr = np.load(path)
        return {
            "format": ext_lower,
            "size_bytes": size,
            "n_rows": arr.shape[0],
            "n_cols": arr.shape[1] if arr.ndim > 1 else 1,
            "dtype": str(arr.dtype),
        }
    elif ext_lower in (".npz",):
        data = np.load(path, allow_pickle=False)
        keys = list(data.keys())
        first = data[keys[0]]
        return {
            "format": ext_lower,
            "size_bytes": size,
            "n_rows": first.shape[0],
            "n_cols": first.shape[1] if first.ndim > 1 else 1,
            "arrays": keys,
        }
    elif ext_lower in {".jpg", ".jpeg", ".png", ".bmp", ".gif",
                       ".webp", ".tiff", ".tif"}:
        try:
            from PIL import Image as PILImage
            img = PILImage.open(path)
            w, h = img.size
            mode = img.mode
            n_channels = len(mode) if mode not in ("L", "P") else 1
            img.close()
            return {
                "format": ext_lower,
                "size_bytes": size,
                "width": w,
                "height": h,
                "n_channels": n_channels,
                "mode": mode,
                "shape": (h, w, n_channels),
            }
        except ImportError:
            return {"format": ext_lower, "size_bytes": size,
                    "note": "Pillow requis pour lire les métadonnées image."}
    elif ext_lower in {".wav", ".flac", ".ogg", ".mp3", ".aiff", ".aif"}:
        info = {"format": ext_lower, "size_bytes": size}
        try:
            import soundfile as sf
            sfinfo = sf.info(path)
            info.update({
                "sample_rate": sfinfo.samplerate,
                "n_channels": sfinfo.channels,
                "n_frames": sfinfo.frames,
                "duration_sec": round(sfinfo.frames / sfinfo.samplerate, 3),
                "subtype": sfinfo.subtype,
            })
            return info
        except ImportError:
            pass
        except Exception:
            pass
        if ext_lower == ".wav":
            try:
                import wave
                with wave.open(path, "rb") as wf:
                    info.update({
                        "sample_rate": wf.getframerate(),
                        "n_channels": wf.getnchannels(),
                        "n_frames": wf.getnframes(),
                        "duration_sec": round(
                            wf.getnframes() / wf.getframerate(), 3),
                        "sample_width": wf.getsampwidth(),
                    })
                return info
            except Exception:
                pass
        info["note"] = "Installez soundfile pour les métadonnées audio complètes."
        return info
    else:
        return {
            "format": ext_lower,
            "size_bytes": size,
            "note": "Utilisez load_dataset() pour charger ce format.",
        }


# ============================================================================
# Images
# ============================================================================

_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".gif", ".webp", ".tiff", ".tif"}


def _load_image_file(path, image_size=None, mode="rgb", normalize=True):
    """Charge une image en numpy array float64.

    Args:
        path       : chemin vers le fichier image.
        image_size : tuple (H, W) pour redimensionner. None = taille originale.
        mode       : "rgb" | "grayscale" | "rgba".
        normalize  : si True, divise par 255 → [0, 1].

    Returns:
        array float64 de shape (H, W, C) ou (H, W) si grayscale.
    """
    try:
        from PIL import Image as PILImage
    except ImportError:
        raise ImportError(
            "Pillow requis pour lire les images. "
            "Installez avec: pip install Pillow"
        )

    pil_mode = {"rgb": "RGB", "grayscale": "L", "rgba": "RGBA"}.get(mode.lower(), "RGB")
    img = PILImage.open(path).convert(pil_mode)

    if image_size is not None:
        img = img.resize((image_size[1], image_size[0]), PILImage.LANCZOS)

    arr = np.array(img, dtype=np.float64)
    if normalize:
        arr = arr / 255.0
    return arr


@_register(_IMAGE_EXTS)
def _load_image(path, target=None, image_size=None, mode="rgb",
                normalize=True, **kw):
    """Charge un fichier image unique.

    Returns:
        (X, None, None) où X est de shape (H, W, C) float64.
    """
    arr = _load_image_file(path, image_size=image_size, mode=mode,
                           normalize=normalize)
    return arr, None, None


def load_image_dataset(path, image_size=None, mode="rgb", normalize=True,
                       extensions=None):
    """Charge un dossier d'images organisé par classe.

    Structure attendue :
        path/
            classe_0/
                img1.jpg
                img2.png
            classe_1/
                img3.jpg

    Args:
        path       : chemin vers le dossier racine.
        image_size : tuple (H, W) pour redimensionner toutes les images.
        mode       : "rgb" | "grayscale" | "rgba".
        normalize  : normaliser vers [0, 1].
        extensions : set d'extensions autorisées. None = _IMAGE_EXTS.

    Returns:
        (X, y, class_names)
        X          : array float64 (n_samples, H, W, C) ou (n_samples, H*W*C) si image_size fourni.
        y          : array int64 (n_samples, 1).
        class_names: liste des noms de classes dans l'ordre des labels.
    """
    if extensions is None:
        extensions = _IMAGE_EXTS

    if not os.path.isdir(path):
        raise NotADirectoryError(f"Pas un dossier : {path}")

    subdirs = sorted([
        d for d in os.listdir(path)
        if os.path.isdir(os.path.join(path, d))
    ])

    if not subdirs:
        # Dossier plat — pas de labels
        files = sorted([
            f for f in os.listdir(path)
            if os.path.splitext(f)[1].lower() in extensions
        ])
        if not files:
            raise ValueError(f"Aucune image trouvée dans : {path}")
        X_list = []
        for fname in files:
            arr = _load_image_file(os.path.join(path, fname),
                                   image_size=image_size, mode=mode,
                                   normalize=normalize)
            X_list.append(arr.flatten() if image_size else arr)
        X = np.array(X_list, dtype=np.float64)
        logger.info("Images (plat) : %d fichiers, X=%s", len(files), X.shape)
        return X, None, []

    class_names = subdirs
    X_list, y_list = [], []

    for label_idx, label in enumerate(class_names):
        label_dir = os.path.join(path, label)
        files = sorted([
            f for f in os.listdir(label_dir)
            if os.path.splitext(f)[1].lower() in extensions
        ])
        for fname in files:
            fpath = os.path.join(label_dir, fname)
            try:
                arr = _load_image_file(fpath, image_size=image_size,
                                       mode=mode, normalize=normalize)
                X_list.append(arr.flatten() if image_size else arr)
                y_list.append(label_idx)
            except Exception as e:
                logger.warning("Image ignorée %s : %s", fpath, e)

    if not X_list:
        raise ValueError(f"Aucune image chargée depuis : {path}")

    X = np.array(X_list, dtype=np.float64)
    y = np.array(y_list, dtype=np.int64).reshape(-1, 1)
    logger.info("Images : %d fichiers, %d classes, X=%s", len(X_list),
                len(class_names), X.shape)
    return X, y, class_names


# ============================================================================
# Audio
# ============================================================================

_AUDIO_EXTS = {".wav", ".flac", ".ogg", ".mp3", ".aiff", ".aif"}


def _load_audio_file(path, sr=None, mono=True, normalize=True,
                     max_len=None):
    """Charge un fichier audio en numpy array float64.

    Priorité : soundfile → pydub → scipy.io.wavfile (wav) → wave stdlib.

    Args:
        path     : chemin vers le fichier audio.
        sr       : sample rate cible (None = original).
        mono     : convertir en mono si True.
        normalize: normaliser vers [-1, 1].
        max_len  : nombre max d'échantillons (tronque ou padde).

    Returns:
        (signal float64 1D, sample_rate int)
    """
    ext = os.path.splitext(path)[1].lower()
    signal = None
    file_sr = None

    # --- soundfile (priorité : supporte flac, ogg, wav) ---------------------
    try:
        import soundfile as sf
        data, file_sr = sf.read(path, always_2d=True)
        if mono:
            data = data.mean(axis=1)
        else:
            data = data.flatten()
        signal = data.astype(np.float64)
    except ImportError:
        pass
    except Exception as e:
        logger.warning("soundfile échoué (%s)", e)

    # --- pydub (mp3 et autres via ffmpeg) -----------------------------------
    if signal is None:
        try:
            from pydub import AudioSegment
            audio = AudioSegment.from_file(path)
            if mono:
                audio = audio.set_channels(1)
            file_sr = audio.frame_rate
            raw = np.array(audio.get_array_of_samples(), dtype=np.float64)
            max_val = float(2 ** (audio.sample_width * 8 - 1))
            signal = raw / max_val
        except ImportError:
            pass
        except Exception as e:
            logger.warning("pydub échoué (%s)", e)

    # --- scipy.io.wavfile (fallback wav) ------------------------------------
    if signal is None and ext == ".wav":
        try:
            from scipy.io import wavfile
            file_sr, data = wavfile.read(path)
            data = data.astype(np.float64)
            if data.ndim > 1 and mono:
                data = data.mean(axis=1)
            max_val = float(np.iinfo(np.int16).max)
            signal = data / max_val
        except ImportError:
            pass
        except Exception as e:
            logger.warning("scipy.io.wavfile échoué (%s)", e)

    # --- wave stdlib (dernier recours, wav uniquement) ----------------------
    if signal is None and ext == ".wav":
        import wave, struct
        with wave.open(path, "rb") as wf:
            n_channels = wf.getnchannels()
            sampwidth = wf.getsampwidth()
            n_frames = wf.getnframes()
            file_sr = wf.getframerate()
            raw = wf.readframes(n_frames)
        fmt = {1: "b", 2: "h", 4: "i"}.get(sampwidth, "h")
        samples = np.array(
            struct.unpack(f"<{n_frames * n_channels}{fmt}", raw),
            dtype=np.float64
        )
        max_val = float(2 ** (8 * sampwidth - 1))
        samples = samples / max_val
        if mono and n_channels > 1:
            samples = samples.reshape(-1, n_channels).mean(axis=1)
        signal = samples

    if signal is None:
        raise RuntimeError(
            f"Impossible de lire {path}. "
            "Installez soundfile (pip install soundfile) "
            "ou pydub (pip install pydub) + ffmpeg."
        )

    # --- Rééchantillonnage --------------------------------------------------
    if sr is not None and file_sr is not None and sr != file_sr:
        ratio = sr / file_sr
        new_len = int(len(signal) * ratio)
        indices = np.linspace(0, len(signal) - 1, new_len)
        signal = np.interp(indices, np.arange(len(signal)), signal)
        file_sr = sr

    # --- Normalisation ------------------------------------------------------
    if normalize:
        peak = np.abs(signal).max()
        if peak > 0:
            signal = signal / peak

    # --- Troncature / padding -----------------------------------------------
    if max_len is not None:
        if len(signal) > max_len:
            signal = signal[:max_len]
        elif len(signal) < max_len:
            signal = np.pad(signal, (0, max_len - len(signal)))

    return signal.astype(np.float64), int(file_sr) if file_sr else 0


@_register(_AUDIO_EXTS)
def _load_audio(path, target=None, sr=None, mono=True, normalize=True,
                max_len=22050, **kw):
    """Charge un fichier audio unique.

    Returns:
        (signal 1D float64, sample_rate) via le tuple (X, y, meta)
        où X = signal, y = None, meta = sample_rate.
    """
    signal, file_sr = _load_audio_file(
        path, sr=sr, mono=mono, normalize=normalize,
        max_len=max_len or kw.get("max_len")
    )
    return signal, None, file_sr


def load_audio_dataset(path, sr=16000, mono=True, normalize=True,
                       max_len=16000, extensions=None):
    """Charge un dossier audio organisé par classe.

    Structure attendue :
        path/
            classe_0/
                clip1.wav
                clip2.flac
            classe_1/
                clip3.wav

    Args:
        path      : chemin vers le dossier racine.
        sr        : sample rate cible (rééchantillonnage si nécessaire).
        mono      : convertir en mono.
        normalize : normaliser vers [-1, 1].
        max_len   : longueur fixe en échantillons (tronque ou padde).
        extensions: set d'extensions autorisées. None = _AUDIO_EXTS.

    Returns:
        (X, y, class_names)
        X          : array float64 (n_samples, max_len).
        y          : array int64 (n_samples, 1).
        class_names: liste des noms de classes.
    """
    if extensions is None:
        extensions = _AUDIO_EXTS

    if not os.path.isdir(path):
        raise NotADirectoryError(f"Pas un dossier : {path}")

    subdirs = sorted([
        d for d in os.listdir(path)
        if os.path.isdir(os.path.join(path, d))
    ])

    if not subdirs:
        # Dossier plat
        files = sorted([
            f for f in os.listdir(path)
            if os.path.splitext(f)[1].lower() in extensions
        ])
        if not files:
            raise ValueError(f"Aucun fichier audio trouvé dans : {path}")
        X_list = []
        for fname in files:
            sig, _ = _load_audio_file(os.path.join(path, fname),
                                      sr=sr, mono=mono, normalize=normalize,
                                      max_len=max_len)
            X_list.append(sig)
        X = np.array(X_list, dtype=np.float64)
        logger.info("Audio (plat) : %d fichiers, X=%s", len(files), X.shape)
        return X, None, []

    class_names = subdirs
    X_list, y_list = [], []

    for label_idx, label in enumerate(class_names):
        label_dir = os.path.join(path, label)
        files = sorted([
            f for f in os.listdir(label_dir)
            if os.path.splitext(f)[1].lower() in extensions
        ])
        for fname in files:
            fpath = os.path.join(label_dir, fname)
            try:
                sig, _ = _load_audio_file(fpath, sr=sr, mono=mono,
                                          normalize=normalize, max_len=max_len)
                X_list.append(sig)
                y_list.append(label_idx)
            except Exception as e:
                logger.warning("Audio ignoré %s : %s", fpath, e)

    if not X_list:
        raise ValueError(f"Aucun fichier audio chargé depuis : {path}")

    X = np.array(X_list, dtype=np.float64)
    y = np.array(y_list, dtype=np.int64).reshape(-1, 1)
    logger.info("Audio : %d fichiers, %d classes, X=%s", len(X_list),
                len(class_names), X.shape)
    return X, y, class_names


# ============================================================================
# load_folder — alias unifié images + audio
# ============================================================================

def load_folder(path, image_size=None, target_size=None, sr=16000,
                max_len=16000, mono=True, normalize=True, extensions=None):
    """Charge un dossier de fichiers (images ou audio) organisés par label.

    Détecte automatiquement si le contenu est des images ou de l'audio.
    Alias pratique pour load_image_dataset / load_audio_dataset.

    Returns:
        (X, y) — class_names non retourné pour compatibilité trainer.py.
    """
    if extensions is None:
        all_exts = _IMAGE_EXTS | _AUDIO_EXTS
    else:
        all_exts = extensions

    # Détecter le type dominant
    sample_files = []
    for root, dirs, files in os.walk(path):
        for f in files:
            ext = os.path.splitext(f)[1].lower()
            if ext in all_exts:
                sample_files.append(ext)
        if len(sample_files) >= 10:
            break

    img_count = sum(1 for e in sample_files if e in _IMAGE_EXTS)
    aud_count = sum(1 for e in sample_files if e in _AUDIO_EXTS)

    if img_count >= aud_count:
        X, y, _ = load_image_dataset(
            path, image_size=image_size or target_size,
            normalize=normalize, extensions=extensions
        )
    else:
        X, y, _ = load_audio_dataset(
            path, sr=sr, mono=mono, normalize=normalize,
            max_len=max_len, extensions=extensions
        )
    return X, y
